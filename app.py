import sqlite3
import io
import openpyxl
from openpyxl.styles import PatternFill
from flask import Flask, render_template, request, redirect, url_for, flash, send_file

app = Flask(__name__)
app.secret_key = 'chave_secreta_super_segura' # Necessário para flash messages
DB_NAME = "restaurante.db"

def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS comanda (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            telefone TEXT,
            numero_comanda TEXT NOT NULL,
            qtd_pessoas INTEGER NOT NULL,
            couvert_lancado INTEGER DEFAULT 0,
            ativa INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

@app.route('/')
def index():
    busca = request.args.get('busca')
    conn = get_db_connection()
    
    query_base = 'SELECT * FROM comanda WHERE ativa = 1'
    params = []
    
    if busca:
        query_base += ' AND numero_comanda LIKE ?'
        params.append(f'%{busca}%')
        
    # Pega comandas ativas filtradas ou todas
    sem_couvert = conn.execute(query_base + ' AND couvert_lancado = 0', params).fetchall()
    com_couvert = conn.execute(query_base + ' AND couvert_lancado = 1', params).fetchall()
    
    conn.close()
    
    return render_template('index.html', sem_couvert=sem_couvert, com_couvert=com_couvert)

@app.route('/adicionar', methods=['POST'])
def adicionar():
    nome = request.form['nome']
    telefone = request.form['telefone']
    numero_comanda = request.form['numero_comanda']
    qtd_pessoas = request.form['qtd_pessoas']
    
    conn = get_db_connection()
    
    # Verifica duplicidade
    existente = conn.execute('SELECT id FROM comanda WHERE numero_comanda = ? AND ativa = 1', (numero_comanda,)).fetchone()
    
    if existente:
        flash(f'Erro: A comanda {numero_comanda} já está aberta!', 'error')
        conn.close()
        return redirect(url_for('index'))
    
    conn.execute('INSERT INTO comanda (nome, telefone, numero_comanda, qtd_pessoas) VALUES (?, ?, ?, ?)',
                 (nome, telefone, numero_comanda, qtd_pessoas))
    conn.commit()
    conn.close()
    
    flash('Comanda adicionada com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/lancar_couvert/<int:id>', methods=['POST'])
def lancar_couvert(id):
    conn = get_db_connection()
    conn.execute('UPDATE comanda SET couvert_lancado = 1 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    flash('Couvert lançado!', 'success')
    return redirect(url_for('index'))

@app.route('/fechar/<int:id>', methods=['POST'])
def fechar(id):
    conn = get_db_connection()
    conn.execute('UPDATE comanda SET ativa = 0 WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    flash('Comanda fechada com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/exportar', methods=['GET'])
def exportar():
    conn = get_db_connection()
    comandas = conn.execute('SELECT * FROM comanda').fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comandas"

    headers = ["ID", "Nome", "Telefone", "Número Comanda", "Qtd Pessoas", "Couvert", "Status", "Criada em"]
    ws.append(headers)

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, comanda in enumerate(comandas, start=2): # Start from row 2 (row 1 is headers)
        row = [
            comanda['id'],
            comanda['nome'],
            comanda['telefone'],
            comanda['numero_comanda'],
            comanda['qtd_pessoas'],
            "Sim" if comanda['couvert_lancado'] else "Não",
            "Ativa" if comanda['ativa'] else "Fechada",
            comanda['created_at']
        ]
        ws.append(row)

        ativa = comanda['ativa']
        couvert_lancado = comanda['couvert_lancado']

        fill_color = white_fill
        if ativa == 0:
            fill_color = green_fill
        elif ativa == 1 and couvert_lancado == 1:
            fill_color = yellow_fill
        else:
            fill_color = white_fill

        for col_num in range(1, len(headers) + 1):
            ws.cell(row=i, column=col_num).fill = fill_color

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='comandas.xlsx'
    )

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000)
